import requests
from bs4 import BeautifulSoup
import openpyxl

count = 0
urlBasic = "https://www.ldoceonline.com/dictionary/"

# File info
FileName = "FILE NAME" # Set file name
Extension = ".xlsx" # Set Extension
Sheet_Num = 0

# Load Excel file
ExcelFileName = FileName + Extension
workbook = openpyxl.load_workbook(ExcelFileName)
Selectworksheet = workbook.sheetnames[Sheet_Num]
worksheet = workbook[Selectworksheet]

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

while True:
    print("- - - - - - - - - - - - - - - - -")
    definition = ""
    count += 1
    counter = 0
    if worksheet.cell(row=count, column=1).value != None:
        urlComplete = urlBasic + worksheet.cell(row=count, column=1).value
        print("Word: " + worksheet.cell(row=count, column=1).value + "\nDefinition:")
        try:
            response = requests.get(urlComplete, headers=headers)
            soup = BeautifulSoup(response.text, 'html.parser')
            span_tags = soup.find_all('span', class_='DEF')
            for tag in span_tags[:2]:
                if counter == 0:
                    definition += tag.get_text()
                else:
                    definition += " / " + tag.get_text()
                counter += 1
                print(tag.get_text())
            worksheet.cell(row=count, column=2).value = definition
        except requests.exceptions.RequestException as e:
            print("Definition not found")
            worksheet.cell(row=count, column=2).value = "Definition not found"
    else:
        break
    
    if count == 1048576:
        break

workbook.save(ExcelFileName)
workbook.close()