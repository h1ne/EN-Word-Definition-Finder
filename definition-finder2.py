import requests
from bs4 import BeautifulSoup
import openpyxl

URL_BASIC = "https://www.ldoceonline.com/dictionary/"

# File info
FILE_PATH = "FILE NAME" # Set file name
EXTENTION = ".xlsx" # Set EXTENTION
SHEET_NUM = 0
MAX_ROWS = 1048576


def fetch_definition(word):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
    }

    complete_url = URL_BASIC + word
    definition = ""
    counter = 0

    try:
        response = requests.get(complete_url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        span_tags = soup.find_all('span', class_='DEF')
        for tag in span_tags[:2]:
            if counter == 0:
                definition += tag.get_text()
            else:
                definition += " / " + tag.get_text()
            counter += 1
            # print(tag.get_text())
        
        return definition
    except requests.exceptions.RequestException as e:
        print("Definition not found")
        return "Definition not found"

def main():
    # Load Excel file
    workbook = openpyxl.load_workbook(FILE_PATH + EXTENTION)
    select_worksheet = workbook.sheetnames[SHEET_NUM]
    worksheet = workbook[select_worksheet]
    
    row_number = 0
    for i in range(MAX_ROWS):
        row_number += 1
        if worksheet.cell(row=row_number, column=1).value is None:
            break
        else:
            word = worksheet.cell(row=row_number, column=1).value
            print("- - - - - - - - - - - - - - - - -")
            print("Word:", word)
            definition = fetch_definition(word)
            print("Definition: ", definition)
            worksheet.cell(row=row_number, column=2).value = definition

    workbook.save(FILE_PATH + EXTENTION)
    workbook.close()

if __name__ == "__main__":
    main()
